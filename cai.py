# coding=utf-8
from pykeyboard import PyKeyboard
from pymouse import PyMouse
import time
import pyHook
import pythoncom
import xlrd
import xlwt
import pyperclip
from pynput import mouse, keyboard
import threading
import sys
import re
from openpyxl import Workbook, load_workbook


def copy():
    k.press_key(k.control_l_key)
    k.tap_key("c")  # 改小写！！！！ 大写的话由于单进程会触发shift键 ctrl键就失效了
    k.release_key(k.control_l_key)


def getCopy(maxTime=1.2):
    # maxTime = 3  # 3秒复制 调用copy() 不管结果对错
    while (maxTime > 0):
        maxTime = maxTime - 0.1
        time.sleep(0.1)
        # print('doing')
        copy()

    result = pyperclip.paste()
    return result


def tapkey(key, count=1, waitTime=0.2):
    for i in range(0, count):
        k.tap_key(key)
        time.sleep(waitTime)
#暂定excel使用


def Quit():
    global end
    end = True




def saveToExcel(code,name,feature):
    caiworkbook = xlrd.open_workbook(caiExcelUrl)
    table_cai = caiworkbook.sheets()[0]
    rowCount = table_cai.nrows

    # if(code=='031003001001'):
    #     print(1)

    if(rowCount!=0):
        for row in  range(rowCount):
            excel_feature = str(table_cai.cell_value(row, 2))
            # if('自动排气阀' in excel_feature):
            #     print(1)
            str1=excel_feature.replace('\r','')
            str1=str1.replace('\n','')
            str2=feature.replace('\r','')
            str2=str2.replace('\n','')

            if(str1==str2):
                print('已存在:')
                print(feature)
                return
    #print('开始存储-------'+feature)
    #saveworkbook = xlrd.open_workbook(saveExcelUrl)
    #wb = excel_copy(saveworkbook)  # 利用xlutils.copy下的copy函数复制
    wb= load_workbook(filename=caiExcelUrl)
    worksheet=wb.active
    worksheet=wb['Sheet1']



    global rowMaxCount
    #print(rowMaxCount)
    worksheet.cell(row=rowMaxCount+1,column=1,value=code)
    worksheet.cell(row=rowMaxCount+1,column=2,value=name)
    worksheet.cell(row=rowMaxCount+1,column=3,value=feature)

    wb.save(caiExcelUrl)
    rowMaxCount=rowMaxCount+1

def Do():

    global start
    if start:
        #dodoododododododoodododododoodododododoododododododododo

        #print(zhucais)
        datas = []
        excel = xlrd.open_workbook(excelUrl)
        table = excel.sheets()[0]
        rowCount = table.nrows
        colCount = table.ncols
        for i in range(rowCount):
            value= str(table.cell_value(i, 4))
            code= str(table.cell_value(i, 4))
            name=str(table.cell_value(i, 5))
            feature=str(table.cell_value(i, 6))

            length=len(code)
            isRightNum=False
            if (str(code).isdigit()):
                if (length == 12):
                    # print('is 12 number')
                    isRightNum = True
            elif(length==13):
                pattern = re.compile('Z\d+')
                match = pattern.match(str(code))
                if (match):
                    isRightNum = True
                    # print('特殊id')
            elif (length == 6):
                pattern = re.compile('\d\dB\d+')
                match = pattern.match(str(code))
                if (match):
                    isRightNum = True
                    #print('特殊id')

            if(isRightNum):

                #这些主材不处理
                IgnoreNames=['套管','刷油'
                    ,'管道支吊架','电视、电话插座',
                             '压力仪表','避雷网'
                             ,'室内消火栓'
                             ,'管道支架']

                # if(name=='管道支吊架'):
                #     print(1)
                needSave=True
                for ignoreName in IgnoreNames:
                    if(ignoreName in name):
                        needSave=False
                        break

                if(needSave):
                    saveToExcel(code,name,feature)

    start=False
    print('ok')






# 我的代码
def onpressed(Key):
    while True:
        # print(Key)
        if (Key == keyboard.Key.caps_lock):  # 开始
            global start
            start = True
            print('go')

        if (Key == keyboard.Key.f3):  # 结束
            sys.exit()

        global end
        if (end):
            sys.exit()
        return True


def main():
    while True:
        # 主程序在这
        Do()


if __name__ == '__main__':
    k = PyKeyboard()
    m = PyMouse()
    end = False
    start = False
    excelUrl = r"C:\Users\Administrator\Desktop\cai\P1-分部分项.xlsx"#to do-------------
    caiExcelUrl=r"C:\Users\Administrator\Desktop\cai\cai.xlsx"#过滤出的名称和特征

    excel = xlrd.open_workbook(excelUrl)
    caiworkbook = xlrd.open_workbook(caiExcelUrl)
    rowMaxCount = caiworkbook.sheets()[0].nrows

    table = excel.sheets()[0]
    rowCount = table.nrows

    threads = []
    t2 = threading.Thread(target=main, args=())
    threads.append(t2)
    for t in threads:
        t.setDaemon(True)
        t.start()
    print('press Capital to start')

    with keyboard.Listener(on_press=onpressed) as listener:
        listener.join()








